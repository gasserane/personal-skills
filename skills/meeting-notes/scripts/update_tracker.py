"""Append a meeting note's actions, decisions and open items to the centralised tracker.

Usage:
    python update_tracker.py --note "<path to Meeting note - ....md>" \
                             --workbook "<path to Meeting actions tracker.xlsx>"

Parses the canonical meeting-note template (see SKILL.md, shared rules):
- '## Next actions at a glance' -> '### <Person>' tables (Deadline | Topic | Action)
- per-topic '### Decisions' numbered items
- per-topic '### Open / parked' bullets

Idempotent: each row carries a stable ID (meeting date + topic + text hash); re-runs
skip IDs already in the workbook. Existing rows are NEVER modified, so Ane's manual
Status / Progress / Notes edits are preserved. The workbook is Ane's working file:
this script only ever appends.

Branding mirrors ane_package.reporting.brand.IPPF_FORMAT_TEMPLATE (self-contained here
so the skill works without the work-folder package on the path).
"""
from __future__ import annotations

import argparse
import hashlib
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Barlow Medium"
DREAM = "00313C"
STATUS_OPTIONS = '"Open,In progress,Blocked,Done,Dropped"'
PENDING_STATUS_OPTIONS = '"Watching,Closed"'

ACTION_HEADERS = ["ID", "Added", "Meeting date", "Counterpart", "Topic", "Action",
                  "Deadline", "Status", "Progress notes", "Last updated"]
DECISION_HEADERS = ["ID", "Meeting date", "Counterpart", "Topic", "Decision", "Notes"]
PENDING_HEADERS = ["ID", "Meeting date", "Counterpart", "Topic", "Item",
                   "Trigger / review by", "Status", "Notes"]
TOPIC_HEADERS = ["Topic", "Counterpart", "Name", "What it is",
                 "My role", "In annual plan", "Strategic importance (IPPF EN)",
                 "Effort for me", "Status", "Recurrence", "Delegation potential",
                 "Energy", "Funding / project", "First seen", "Last discussed",
                 "Times discussed", "Notes"]
# Topics judgement-column drop-downs: column letter -> options
TOPIC_DROPDOWNS = {
    "E": '"Lead,Co-lead,Co-worker,Advisor"',
    "G": '"High,Medium,Low"',
    "H": '"High,Medium,Low"',
    "I": '"Active,Dormant,Closed"',
    "J": '"Recurring,One-off"',
    "K": '"High,Medium,Low"',
    "L": '"Gives,Neutral,Drains"',
}
COL_WIDTHS = {"ID": 20, "Added": 11, "Meeting date": 12, "Counterpart": 12,
              "Topic": 8, "Action": 70, "Decision": 80, "Item": 70,
              "Deadline": 24, "Trigger / review by": 30, "Status": 12,
              "Progress notes": 40, "Notes": 40, "Last updated": 12,
              "Name": 38, "What it is": 55, "My role": 11, "In annual plan": 26,
              "Strategic importance (IPPF EN)": 14, "Effort for me": 12,
              "Recurrence": 11, "Delegation potential": 12, "Energy": 9,
              "Funding / project": 16, "First seen": 11, "Last discussed": 12,
              "Times discussed": 10}

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}


def parse_meeting_date(text: str) -> str:
    m = re.search(r"\*\*Date:\*\*\s*(\d{1,2})\s+(\w+)\s+(\d{4})", text)
    if m and m.group(2).lower() in MONTHS:
        return f"{int(m.group(3)):04d}-{MONTHS[m.group(2).lower()]:02d}-{int(m.group(1)):02d}"
    m = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    return m.group(1) if m else date.today().isoformat()


def clean(s: str) -> str:
    return re.sub(r"\*\*", "", s).strip()


def make_id(meeting_date: str, topic: str, text: str) -> str:
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:6]
    return f"{meeting_date}-{topic or 'GEN'}-{digest}"


def parse_note(path: Path):
    """Return (meeting_date, counterpart, actions, decisions, pending).

    actions: list of (person, topic, action, deadline)
    decisions: list of (topic, decision)
    pending: list of (topic, item, trigger)
    """
    text = path.read_text(encoding="utf-8")
    meeting_date = parse_meeting_date(text)
    counterpart = ""
    m = re.search(r"\*\*Participants:\*\*\s*(.+)", text)
    if m:
        names = [n.strip() for n in m.group(1).split(",")]
        others = [n for n in names if not n.lower().startswith("ane")]
        counterpart = others[0].split()[0] if others else ""

    actions, decisions, pending = [], [], []
    topics = {}  # tag -> full name
    current_h2, current_h3, current_topic = "", "", ""

    for raw in text.splitlines():
        line = raw.rstrip()
        s = line.strip()
        if s.startswith("## "):
            current_h2 = clean(s[3:])
            current_h3 = ""
            tm = re.match(r"(T\d+)\b", current_h2)
            current_topic = tm.group(1) if tm else ""
            if current_topic:
                name = re.sub(r"^T\d+\s*[-:]\s*", "", current_h2).strip()
                topics[current_topic] = name
            continue
        if s.startswith("### "):
            current_h3 = clean(s[4:])
            continue
        in_actions_block = current_h2.lower().startswith("next actions")
        if in_actions_block and s.startswith("|"):
            cells = [clean(c) for c in s.strip("|").split("|")]
            if len(cells) >= 3 and cells[0].lower() != "deadline" \
                    and not all(set(c) <= set(":- ") for c in cells):
                person = re.sub(r"\s*\(.*\)$", "", current_h3).strip() or "Ane"
                deadline, topic, action = cells[0], cells[1], cells[2]
                if action and action != "-":
                    actions.append((person, topic, action, deadline))
            continue
        if current_topic and current_h3.lower().startswith("decisions"):
            dm = re.match(r"\d+\.\s+(.*)", s)
            if dm:
                decisions.append((current_topic, clean(dm.group(1))))
            continue
        if current_topic and current_h3.lower().startswith("open"):
            if s.startswith("- "):
                item = clean(s[2:])
                trig = ""
                tm = re.search(r"\(([^)]*(?:pending|by |after |due |revisit|confirm)[^)]*)\)",
                               item, re.IGNORECASE)
                if tm:
                    trig = tm.group(1)
                pending.append((current_topic, item, trig))
            continue

    return meeting_date, counterpart, actions, decisions, pending, topics


def style_header(ws, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DREAM)
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(h, 16)
    ws.freeze_panes = "A2"


def ensure_sheet(wb, name, headers, status_col=None, status_options=STATUS_OPTIONS):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    style_header(ws, headers)
    if status_col:
        dv = DataValidation(type="list", formula1=status_options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{status_col}2:{status_col}1000")
    return ws


def ensure_topics_sheet(wb):
    if "Topics" in wb.sheetnames:
        return wb["Topics"]
    ws = wb.create_sheet("Topics", 1)  # right after Guide
    style_header(ws, TOPIC_HEADERS)
    for col, options in TOPIC_DROPDOWNS.items():
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")
    return ws


def upsert_topics(ws, topics, meeting_date, counterpart):
    """Script-owned cells only: Topic, Counterpart, Name, First seen, Last discussed,
    Times discussed. Judgement columns (role, plan link, importance, effort, status,
    recurrence, delegation, energy, funding, notes) belong to Ane and are never
    written for existing rows. Idempotent: same meeting date never double-counts."""
    added = updated = 0
    index = {}
    for r in range(2, ws.max_row + 1):
        tag, cp = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if tag:
            index[(tag, cp or "")] = r
    for tag, name in sorted(topics.items()):
        key = (tag, counterpart or "")
        if key in index:
            r = index[key]
            last = str(ws.cell(row=r, column=15).value or "")
            if meeting_date > last:
                ws.cell(row=r, column=15, value=meeting_date)
                ws.cell(row=r, column=16,
                        value=int(ws.cell(row=r, column=16).value or 0) + 1)
                updated += 1
            if not ws.cell(row=r, column=3).value:
                ws.cell(row=r, column=3, value=name)
        else:
            append_row(ws, [tag, counterpart, name, "", "", "", "", "", "Active",
                            "", "", "", "", meeting_date, meeting_date, 1, ""])
            added += 1
    return added, updated


def existing_ids(ws):
    return {ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}


def append_row(ws, values):
    r = ws.max_row + 1
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)


GUIDE_LINES = [
    ("Meeting actions tracker - how to use this workbook", True),
    ("", False),
    ("What this is: one central place for every action, decision and pending item "
     "from your meeting notes, added automatically when a note is confirmed.", False),
    ("", False),
    ("Tabs: 'Topics' is the registry of what each topic tag means, with your own "
     "metadata per topic. One tab per person holds that person's actions (your tab is "
     "what you owe; another person's tab is what you chase with them). 'Decisions' is "
     "the decision log. 'Pending' holds open and parked items with their resurface "
     "trigger.", False),
    ("", False),
    ("Yours to edit: Status, Progress notes and Notes columns on the action tabs, and "
     "on Topics all the judgement columns: My role (Lead / Co-lead / Co-worker / "
     "Advisor), In annual plan (point to the subgoal, e.g. 'Yes - 1.2' or 'New item'), "
     "Strategic importance for IPPF EN, Effort for me, Status, Recurrence, Delegation "
     "potential, Energy (does the topic give or drain energy), Funding / project. The "
     "updater never changes cells you own, so your edits are safe.", False),
    ("Added automatically: everything else. Action, decision and pending rows carry a "
     "stable ID so re-running the updater never duplicates. On Topics the updater "
     "maintains First seen, Last discussed and Times discussed per topic.", False),
    ("", False),
    ("Why the Topics metadata: at year end, filter high Effort + low Strategic "
     "importance (delegate or drop candidates), compare In annual plan vs New item "
     "(how much of your agenda was planned work), and read Energy against Times "
     "discussed. This is the dataset for optimising next year's planning.", False),
    ("", False),
    ("Status meanings: Open = not started. In progress = started. Blocked = waiting "
     "on someone or something. Done = finished. Dropped = agreed not to do.", False),
    ("Pending status: Watching = still to monitor. Closed = resolved or absorbed "
     "into an action.", False),
]


def ensure_guide(wb):
    if "Guide" in wb.sheetnames:
        return
    ws = wb.create_sheet("Guide", 0)
    ws.column_dimensions["A"].width = 110
    for i, (line, bold) in enumerate(GUIDE_LINES, start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT, size=12 if bold else 10, bold=bold,
                      color=DREAM if bold else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--note", required=True)
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    note = Path(args.note)
    book = Path(args.workbook)
    meeting_date, counterpart, actions, decisions, pending, topics = parse_note(note)
    today = date.today().isoformat()

    wb = load_workbook(book) if book.exists() else Workbook()
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        del wb["Sheet"]
    ensure_guide(wb)
    t_added, t_updated = upsert_topics(ensure_topics_sheet(wb), topics,
                                       meeting_date, counterpart)

    added = {"actions": 0, "decisions": 0, "pending": 0}

    for person, topic, action, deadline in actions:
        ws = ensure_sheet(wb, person, ACTION_HEADERS, status_col="H")
        rid = make_id(meeting_date, topic, person + action)
        if rid in existing_ids(ws):
            continue
        append_row(ws, [rid, today, meeting_date, counterpart, topic, action,
                        deadline, "Open", "", today])
        added["actions"] += 1

    ws_d = ensure_sheet(wb, "Decisions", DECISION_HEADERS)
    ids_d = existing_ids(ws_d)
    for topic, decision in decisions:
        rid = make_id(meeting_date, topic, decision)
        if rid in ids_d:
            continue
        append_row(ws_d, [rid, meeting_date, counterpart, topic, decision, ""])
        added["decisions"] += 1

    ws_p = ensure_sheet(wb, "Pending", PENDING_HEADERS, status_col="G",
                        status_options=PENDING_STATUS_OPTIONS)
    ids_p = existing_ids(ws_p)
    for topic, item, trig in pending:
        rid = make_id(meeting_date, topic, item)
        if rid in ids_p:
            continue
        append_row(ws_p, [rid, meeting_date, counterpart, topic, item, trig,
                          "Watching", ""])
        added["pending"] += 1

    wb.save(book)
    print(f"Tracker updated: {book}")
    print(f"  note: {note.name} (meeting {meeting_date}, counterpart {counterpart or '?'})")
    print(f"  added {added['actions']} actions, {added['decisions']} decisions, "
          f"{added['pending']} pending items (existing rows untouched)")
    print(f"  topics: {t_added} new, {t_updated} last-discussed updated "
          f"(your judgement columns untouched)")


if __name__ == "__main__":
    main()
